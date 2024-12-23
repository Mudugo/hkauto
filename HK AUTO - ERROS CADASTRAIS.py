import pyautogui
import time
import os
from datetime import datetime
from openpyxl import load_workbook
import pywhatkit as kit
import psutil

def abrir_sarclient():
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    sarclient_path = os.path.join(desktop, "SarClient.lnk")
    os.startfile(sarclient_path)

def login():
    image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\login.png"
    login_button = None
    timeout = time.time() + 600
    
    while login_button is None and time.time() < timeout:
        try:
            login_button = pyautogui.locateOnScreen(image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if login_button is not None:
        x, y = pyautogui.center(login_button)
        pyautogui.click(x, y)
        pyautogui.press('delete')
        pyautogui.click(x, y)
        pyautogui.write('xxxx')
        pyautogui.press('tab')
        pyautogui.write('xxxx')
        pyautogui.press('enter')
    else:
        main()

def acessar_modulo():
    modulo_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\modulo.png"
    modulo_button = None
    timeout = time.time() + 600
    
    while modulo_button is None and time.time() < timeout:
        try:
            modulo_button = pyautogui.locateOnScreen(modulo_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if modulo_button is not None:
        x_mod, y_mod = pyautogui.center(modulo_button)
        pyautogui.click(x_mod, y_mod)
    else:
        main()

def acessar_modulo_relatório():
    relatorio_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\relatorio.png"
    relatorio_button = None
    timeout = time.time() + 600
    
    while relatorio_button is None and time.time() < timeout:
        try:
            relatorio_button = pyautogui.locateOnScreen(relatorio_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if relatorio_button is not None:
        x_mod, y_mod = pyautogui.center(relatorio_button)
        pyautogui.click(x_mod, y_mod)
    else:
        main()

def acessar_rh():
    rh_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\rh.png"
    rh_button = None
    timeout = time.time() + 600
    
    while rh_button is None and time.time() < timeout:
        try:
            rh_button = pyautogui.locateOnScreen(rh_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if rh_button is not None:
        x_mod, y_mod = pyautogui.center(rh_button)
        pyautogui.doubleClick(x_mod, y_mod)
    else:
        main()

def acessar_dadoscadastrais():
    dadoscadastrais_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\dadoscadastrais.png"
    dadoscadastrais_button = None
    timeout = time.time() + 600
    
    while dadoscadastrais_button is None and time.time() < timeout:
        try:
            dadoscadastrais_button = pyautogui.locateOnScreen(dadoscadastrais_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if dadoscadastrais_button is not None:
        x_mod, y_mod = pyautogui.center(dadoscadastrais_button)
        pyautogui.doubleClick(x_mod, y_mod)
    else:
        main()

def acessar_situacao():
    situacao_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\situacao.png"
    situacao_button = None
    timeout = time.time() + 600
    
    while situacao_button is None and time.time() < timeout:
        try:
            situacao_button = pyautogui.locateOnScreen(situacao_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if situacao_button is not None:
        x_mod, y_mod = pyautogui.center(situacao_button)
        pyautogui.click(x_mod + 70, y_mod)
        pyautogui.press('tab')
        pyautogui.write('ATIVO')
    else:
        main()

def exportar1():
    exportar1_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\exportar1.png"
    exportar1_button = None
    timeout = time.time() + 600
    
    while exportar1_button is None and time.time() < timeout:
        try:
            exportar1_button = pyautogui.locateOnScreen(exportar1_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if exportar1_button is not None:
        x_mod, y_mod = pyautogui.center(exportar1_button)
        pyautogui.click(x_mod, y_mod)
    else:
        main()

def exportar2():
    exportar2_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\exportar2.png"
    exportar2_button = None
    timeout = time.time() + 600
    
    while exportar2_button is None and time.time() < timeout:
        try:
            exportar2_button = pyautogui.locateOnScreen(exportar2_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if exportar2_button is not None:
        x_mod, y_mod = pyautogui.center(exportar2_button)
        pyautogui.click(x_mod, y_mod)
    else:
        main()

def inserir_nome(data_hora):
    nome_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\nome.png"
    nome_button = None
    timeout = time.time() + 600
    
    while nome_button is None and time.time() < timeout:
        try:
            nome_button = pyautogui.locateOnScreen(nome_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if nome_button is not None:
        x_mod, y_mod = pyautogui.center(nome_button)
        pyautogui.click(x_mod, y_mod)
        pyautogui.write(data_hora)
    else:
        main()
        
def area_trabalho():
    areatrabalho_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\areatrabalho.png"
    areatrabalho_button = None
    timeout = time.time() + 600
    
    while areatrabalho_button is None and time.time() < timeout:
        try:
            areatrabalho_button = pyautogui.locateOnScreen(areatrabalho_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if areatrabalho_button is not None:
        x_mod, y_mod = pyautogui.center(areatrabalho_button)
        pyautogui.click(x_mod, y_mod)
    else:
        main()

def salvar():
    salvar_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\salvar.png"
    salvar_button = None
    timeout = time.time() + 600
    
    while salvar_button is None and time.time() < timeout:
        try:
            salvar_button = pyautogui.locateOnScreen(salvar_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if salvar_button is not None:
        x_mod, y_mod = pyautogui.center(salvar_button)
        pyautogui.click(x_mod, y_mod)
    else:
        main()

def ok():
    ok_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\ok.png"
    ok_button = None
    timeout = time.time() + 600
    
    while ok_button is None and time.time() < timeout:
        try:
            ok_button = pyautogui.locateOnScreen(ok_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if ok_button is not None:
        x_mod, y_mod = pyautogui.center(ok_button)
        pyautogui.click(x_mod, y_mod)
    else:
        main()

def arquivo(data_hora):
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    arquivo_excel = os.path.join(desktop, f"{data_hora}.xlsx")
    return arquivo_excel
    
def aplicar_limpa(arquivo_excel):
    wb = load_workbook(arquivo_excel)
    ws = wb.active

    colunas_manter = [1, 3, 4, 5, 6, 8, 35, 36, 37, 38, 39]

    for col in range(ws.max_column, 0, -1):
        if col not in colunas_manter:
            ws.delete_cols(col)
    
    wb.save(arquivo_excel)
    wb.close()

def aplicar_formatacao_condicional(arquivo_excel):
    wb = load_workbook(arquivo_excel)
    ws = wb.active

    col_re = 2
    col_nome = 3
    colunas_acento = [col_nome, 4, 7, 8, 9]
    col_senha = 5
    col_genero = 6

    re_values = {}
    
    erros_re_duplicado = []
    erros_acento_nome = []
    erros_acento_apelido = []
    erros_acento_nome_mae = []
    erros_acento_nome_pai = []
    erros_acento_nome_conjuge = []
    erros_senhor_senhora = []
    erros_espaco = []

    for row in range(2, ws.max_row + 1):
        re_value = ws.cell(row=row, column=col_re).value
        nome_value = ws.cell(row=row, column=col_nome).value

        for col in [col_nome, 4, 7, 8, 9]:
            cell_value = ws.cell(row=row, column=col).value
            if isinstance(cell_value, str):
                if cell_value != cell_value.strip() or '  ' in cell_value:
                    erros_espaco.append(row)
                    break

        if re_value:
            if re_value in re_values:
                re_values[re_value].append((row, nome_value))
            else:
                re_values[re_value] = [(row, nome_value)]

    for re_value, values in re_values.items():
        if len(values) > 1:
            nomes = set(nome for _, nome in values)
            if len(nomes) > 1:
                erros_re_duplicado.extend(row for row, _ in values)

    for row in range(2, ws.max_row + 1):
        acentuacao_found = False

        for col in colunas_acento:
            cell_value = ws.cell(row=row, column=col).value
            if cell_value == "NÃO CADASTRADO":
                continue
            if cell_value and any(char in cell_value for char in ['á', 'é', 'í', 'ó', 'ú', 'â', 'ê', 'ô', 'ã', 'õ', 'ü', 'ç', 'Á', 'É', 'Í', 'Ó', 'Ú', 'Â', 'Ê', 'Ô', 'Ã', 'Õ', 'Ü', 'Ç']):
                acentuacao_found = True
                if col == col_nome:
                    erros_acento_nome.append(row)
                elif col == 4:
                    erros_acento_apelido.append(row)
                elif col == 7:
                    erros_acento_nome_mae.append(row)
                elif col == 8:
                    erros_acento_nome_pai.append(row)
                elif col == 9:
                    erros_acento_nome_conjuge.append(row)
                break

        senhor_value = ws.cell(row=row, column=col_senha).value
        genero_value = ws.cell(row=row, column=col_genero).value
        valid_comparisons = not (
            (senhor_value == 'SENHOR' and genero_value != 'M') or 
            (senhor_value == 'SENHORA' and genero_value != 'F')
        )

        if not acentuacao_found and not valid_comparisons:
            erros_senhor_senhora.append(row)

    erros_encontrados = any([
        erros_re_duplicado, erros_acento_nome, erros_acento_apelido,
        erros_acento_nome_mae, erros_acento_nome_pai, erros_acento_nome_conjuge,
        erros_senhor_senhora, erros_espaco
    ])

    for erro, rows in [
        ('Erros RE Duplicado', erros_re_duplicado),
        ('Erros Nome', erros_acento_nome),
        ('Erros Apelido', erros_acento_apelido),
        ('Erros Nome Mãe', erros_acento_nome_mae),
        ('Erros Nome Pai', erros_acento_nome_pai),
        ('Erros Nome Cônjuge', erros_acento_nome_conjuge),
        ('Erros Senhor Senhora', erros_senhor_senhora),
        ('Erros Espaços', erros_espaco)
    ]:
        if rows:
            ws_erros = wb.create_sheet(title=erro)
            
            for col in range(1, ws.max_column + 1):
                ws_erros.cell(row=1, column=col).value = ws.cell(row=1, column=col).value

            for r in range(len(rows)):
                for c in range(1, ws.max_column + 1):
                    ws_erros.cell(row=r + 2, column=c).value = ws.cell(row=rows[r], column=c).value

    wb.save(arquivo_excel)
    wb.close()

    return erros_encontrados


def enviar_mensagem_whatsapp(arquivo_excel, erros_encontrados):
    numero_destino = "+5511984450333"

    agora = datetime.now()
    data_atual = agora.strftime("%d/%m/%Y")

    if erros_encontrados:
        mensagem = f"Inconsistências encontradas na verificação de {data_atual}."
        kit.sendwhatmsg_instantly(numero_destino, mensagem, 5)
        time.sleep(2)
        anexarwpp_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\ANEXARWPP.png"
        anexarwpp_button = None
        timeout = time.time() + 600
        
        while anexarwpp_button is None and time.time() < timeout:
            try:
                anexarwpp_button = pyautogui.locateOnScreen(anexarwpp_image_path, confidence=0.8)
            except pyautogui.ImageNotFoundException:
                pass
            time.sleep(1)

        if anexarwpp_button is not None:
            x_mod, y_mod = pyautogui.center(anexarwpp_button)
            pyautogui.click(x_mod, y_mod)
        time.sleep(1)
        
        anexardocumento_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\ANEXARDOCUMENTO.png"
        anexardocumento_button = None
        timeout = time.time() + 600
        
        while anexardocumento_button is None and time.time() < timeout:
            try:
                anexardocumento_button = pyautogui.locateOnScreen(anexardocumento_image_path, confidence=0.8)
            except pyautogui.ImageNotFoundException:
                pass
            time.sleep(1)

        if anexardocumento_button is not None:
            x_mod, y_mod = pyautogui.center(anexardocumento_button)
            pyautogui.click(x_mod, y_mod)
        time.sleep(1)
        
        pyautogui.write(arquivo_excel)
        pyautogui.press('enter')
        time.sleep(3)
        pyautogui.press('enter')
    else:
        mensagem = f"Não foram encontradas inconsistências na verificação de {data_atual}."
        kit.sendwhatmsg_instantly(numero_destino, mensagem, 15)
    
def fechar_sarclient():
    for proc in psutil.process_iter(['name']):
        if proc.info['name'] == 'SarClient.exe':
            os.system("taskkill /f /im SarClient.exe")
            return
    pass

def main():
    agora = datetime.now()
    data_hora = agora.strftime("%Y%m%d%H%M")

    fechar_sarclient()
    abrir_sarclient()
    login()
    acessar_modulo()
    acessar_modulo_relatório()
    acessar_rh()
    acessar_dadoscadastrais()
    acessar_situacao()
    exportar1()
    exportar2()
    area_trabalho()
    inserir_nome(data_hora)
    salvar()
    ok()
    fechar_sarclient()
    arquivo_excel = arquivo(data_hora)
    aplicar_limpa(arquivo_excel)
    erros_encontrados = aplicar_formatacao_condicional(arquivo_excel)
    enviar_mensagem_whatsapp(arquivo_excel, erros_encontrados)

if __name__ == "__main__":
    main()
    