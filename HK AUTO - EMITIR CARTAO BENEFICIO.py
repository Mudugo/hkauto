import pyautogui
import time
import os
from datetime import datetime
from openpyxl import load_workbook
import pywhatkit as kit
import psutil

def abrir_sarclient():
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    sarclient_path = os.path.join(desktop, "SarClient.lnk")
    os.startfile(sarclient_path)

def login():
    image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\login.png"
    login_button = None
    timeout = time.time() + 600
    
    while login_button is None and time.time() < timeout:
        try:
            login_button = pyautogui.locateOnScreen(image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if login_button is not None:
        x, y = pyautogui.center(login_button)
        pyautogui.click(x, y)
        pyautogui.press('delete')
        pyautogui.click(x, y)
        pyautogui.write('xxxx')
        pyautogui.press('tab')
        pyautogui.write('xxxx')
        pyautogui.press('enter')
    else:
        main()

def acessar_modulo():
    modulo_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\modulo.png"
    modulo_button = None
    timeout = time.time() + 600
    
    while modulo_button is None and time.time() < timeout:
        try:
            modulo_button = pyautogui.locateOnScreen(modulo_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if modulo_button is not None:
        x_mod, y_mod = pyautogui.center(modulo_button)
        pyautogui.click(x_mod, y_mod)
    else:
        main()

def acessar_modulo_relatório():
    relatorio_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\relatorio.png"
    relatorio_button = None
    timeout = time.time() + 600
    
    while relatorio_button is None and time.time() < timeout:
        try:
            relatorio_button = pyautogui.locateOnScreen(relatorio_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if relatorio_button is not None:
        x_mod, y_mod = pyautogui.center(relatorio_button)
        pyautogui.click(x_mod, y_mod)
    else:
        main()

def acessar_rh():
    rh_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\rh.png"
    rh_button = None
    timeout = time.time() + 600
    
    while rh_button is None and time.time() < timeout:
        try:
            rh_button = pyautogui.locateOnScreen(rh_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if rh_button is not None:
        x_mod, y_mod = pyautogui.center(rh_button)
        pyautogui.doubleClick(x_mod, y_mod)
    else:
        main()

def acessar_dadoscadastrais():
    dadoscadastrais_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\dadoscadastrais.png"
    dadoscadastrais_button = None
    timeout = time.time() + 600
    
    while dadoscadastrais_button is None and time.time() < timeout:
        try:
            dadoscadastrais_button = pyautogui.locateOnScreen(dadoscadastrais_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if dadoscadastrais_button is not None:
        x_mod, y_mod = pyautogui.center(dadoscadastrais_button)
        pyautogui.doubleClick(x_mod, y_mod)
    else:
        main()

def acessar_situacao():
    situacao_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\situacao.png"
    situacao_button = None
    timeout = time.time() + 600
    
    while situacao_button is None and time.time() < timeout:
        try:
            situacao_button = pyautogui.locateOnScreen(situacao_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if situacao_button is not None:
        x_mod, y_mod = pyautogui.center(situacao_button)
        pyautogui.click(x_mod + 70, y_mod)
        pyautogui.press('tab')
        pyautogui.write('ATIVO')
    else:
        main()

def exportar1():
    exportar1_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\exportar1.png"
    exportar1_button = None
    timeout = time.time() + 600
    
    while exportar1_button is None and time.time() < timeout:
        try:
            exportar1_button = pyautogui.locateOnScreen(exportar1_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if exportar1_button is not None:
        x_mod, y_mod = pyautogui.center(exportar1_button)
        pyautogui.click(x_mod, y_mod)
    else:
        main()

def exportar2():
    exportar2_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\exportar2.png"
    exportar2_button = None
    timeout = time.time() + 600
    
    while exportar2_button is None and time.time() < timeout:
        try:
            exportar2_button = pyautogui.locateOnScreen(exportar2_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if exportar2_button is not None:
        x_mod, y_mod = pyautogui.center(exportar2_button)
        pyautogui.click(x_mod, y_mod)
    else:
        main()

def inserir_nome(data_hora):
    nome_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\nome.png"
    nome_button = None
    timeout = time.time() + 600
    
    while nome_button is None and time.time() < timeout:
        try:
            nome_button = pyautogui.locateOnScreen(nome_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if nome_button is not None:
        x_mod, y_mod = pyautogui.center(nome_button)
        pyautogui.click(x_mod, y_mod)
        pyautogui.write(data_hora)
    else:
        main()
        
def area_trabalho():
    areatrabalho_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\areatrabalho.png"
    areatrabalho_button = None
    timeout = time.time() + 600
    
    while areatrabalho_button is None and time.time() < timeout:
        try:
            areatrabalho_button = pyautogui.locateOnScreen(areatrabalho_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if areatrabalho_button is not None:
        x_mod, y_mod = pyautogui.center(areatrabalho_button)
        pyautogui.click(x_mod, y_mod)
    else:
        main()

def salvar():
    salvar_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\salvar.png"
    salvar_button = None
    timeout = time.time() + 600
    
    while salvar_button is None and time.time() < timeout:
        try:
            salvar_button = pyautogui.locateOnScreen(salvar_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if salvar_button is not None:
        x_mod, y_mod = pyautogui.center(salvar_button)
        pyautogui.click(x_mod, y_mod)
    else:
        main()

def ok():
    ok_image_path = r"C:\Users\isaacholanda\Desktop\HK AUTO\capturas\ok.png"
    ok_button = None
    timeout = time.time() + 600
    
    while ok_button is None and time.time() < timeout:
        try:
            ok_button = pyautogui.locateOnScreen(ok_image_path, confidence=0.8)
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(1)

    if ok_button is not None:
        x_mod, y_mod = pyautogui.center(ok_button)
        pyautogui.click(x_mod, y_mod)
    else:
        main()

def arquivo(data_hora):
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    arquivo_excel = os.path.join(desktop, f"{data_hora}.xlsx")
    return arquivo_excel
    
def aplicar_limpa(arquivo_excel):
    wb = load_workbook(arquivo_excel)
    ws = wb.active

    colunas_manter = [1, 4, 9, 17]

    for col in range(ws.max_column, 0, -1):
        if col not in colunas_manter:
            ws.delete_cols(col)
    
    wb.save(arquivo_excel)
    wb.close()

def adicionar_coluna_codigo(arquivo_excel):
    wb = load_workbook(arquivo_excel)
    ws = wb.active
    
    # Adicionar nova coluna após a coluna do CPF (assumindo que a coluna CPF é a quarta)
    coluna_nova_index = 5
    ws.insert_cols(coluna_nova_index)
    ws.cell(row=1, column=coluna_nova_index).value = "Código"

    # Atribuir valores na nova coluna com base na primeira coluna
def adicionar_codigo_empresa(arquivo_excel):
    """
    Adiciona uma nova coluna ao lado da coluna CPF com base nos valores da primeira coluna.
    """
    wb = load_workbook(arquivo_excel)
    ws = wb.active
    
    # Definir o mapeamento entre os valores da coluna 1 e os códigos
    mapeamento = {
        "BRANDONE SERVICE - BRD1": "041",
        "BRANDONE SERVICE - BRD2": "042",
        "BRANDONE SERVICE - BRD3": "043",
        "BRANDONE SERVICE - BRD4": "044",
        "BRANDONE SERVICE - BRD5": "045",
        "BRANDONE SERVICE - UNIT1": "031",
        "BRANDONE SERVICE - UNIT2": "032",
        "BRANDONE SERVICE - UNIT3": "033",
        "BRANDONE SERVICE - UNIT4": "034",
        "BRANDONE SERVICE - UNIT5": "035",
        "BRANDONE SERVICE - UNIT6": "036",
        "BRANDONE SERVICE - UNIT7": "037",
        "BRANDONE SERVICE - UNIT8": "038",
        "GAPE SERVIÇOS": "051",
        "GAPESERV SEGURANÇA": "053",
        "TRS SEGURANÇA": "054",
    }
    
    # Adicionar nova coluna após a coluna do CPF (assumindo que a coluna CPF é a quarta)
    coluna_nova_index = 5
    ws.insert_cols(coluna_nova_index)
    ws.cell(row=1, column=coluna_nova_index).value = "Código"

    # Atribuir valores na nova coluna com base no mapeamento
    for row in range(2, ws.max_row + 1):
        valor_coluna1 = ws.cell(row=row, column=1).value
        ws.cell(row=row, column=coluna_nova_index).value = mapeamento.get(valor_coluna1, "N/A")  # "N/A" para valores não encontrados

    # Salvar alterações
    wb.save(arquivo_excel)
    wb.close()

 
def fechar_sarclient():
    for proc in psutil.process_iter(['name']):
        if proc.info['name'] == 'SarClient.exe':
            os.system("taskkill /f /im SarClient.exe")
            return
    pass

def main():
    agora = datetime.now()
    data_hora = agora.strftime("%Y%m%d%H%M")

    fechar_sarclient()
    abrir_sarclient()
    login()
    acessar_modulo()
    acessar_modulo_relatório()
    acessar_rh()
    acessar_dadoscadastrais()
    acessar_situacao()
    exportar1()
    exportar2()
    area_trabalho()
    inserir_nome(data_hora)
    salvar()
    ok()
    fechar_sarclient()
    arquivo_excel = arquivo(data_hora)
    aplicar_limpa(arquivo_excel)
    adicionar_coluna_codigo(arquivo_excel)
    adicionar_codigo_empresa(arquivo_excel)

if __name__ == "__main__":
    main()
    